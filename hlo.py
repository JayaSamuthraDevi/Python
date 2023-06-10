# import pandas lib as pd
import pandas as pd
 
# read by default 1st sheet of an excel file
dataframe1 = pd.read_excel('book2.xlsx')
 
print(dataframe1)



"""
import openpyxl
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("Book2.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values


for row in range(0, dataframe1.max_row):
    c=0
    z=1
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        c=col[row].value+c
        print(z,col[row].value)
        z=z+1
    
    print(c)
    
"""